import inspect
import logging
import softest
from openpyxl import Workbook, load_workbook
import csv

class Utils(softest.TestCase):
    def assertListItemText(self, list, value):
        for stop in list:
            print("The text is: "+ stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()

    def cust_logger(loglevel=logging.DEBUG):
        # Set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # create logger. To set the level
        logger = logging.getLogger(logger_name)
        logger.setLevel(logging.DEBUG)
        # create file handler
        fh = logging.FileHandler("automation.log")
        # create formatter - how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s -%(name)s : %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # log to the file
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct+1):
            row = []
            for j in range(1, col_ct+1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return(datalist)

    def read_data_from_csv(filename):
        # Create an empty list
        datalist = []
        # Open csv file
        csvdata = open(filename, "r")
        #Create CSV reader
        reader = csv.reader(csvdata)
        # Skip reader. The header will be skipped
        next(reader)
        # Add csv rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist

